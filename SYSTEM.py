import os
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ================= الإعدادات =================
GROUPS = ["4ب", "5ب", "6ب", "1إعدادي", "2إعدادي", "3إعدادي", "1ثانوي", "2ثانوي", "3ثانوي"]
MONTHS = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو", "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]


class DataEngine:
    def __init__(self, group_name):
        self.group = group_name
        if not os.path.exists(self.group): os.makedirs(self.group)
        self.db_path = os.path.join(self.group, "system_data.db")
        self.path_info = os.path.join(self.group, f"أسماء_طلاب_{group_name}.xlsx")
        self.path_pay = os.path.join(self.group, f"سجل_مدفوعات_{group_name}.xlsx")
        self.setup_db()

    def setup_db(self):
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS students (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT)")
            pay_cols = ", ".join([f"[{m}] TEXT DEFAULT 'X'" for m in MONTHS])
            cur.execute(f"CREATE TABLE IF NOT EXISTS payments (sid INTEGER PRIMARY KEY, {pay_cols})")
            conn.commit()

    def auto_save_all(self):
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT id, name FROM students ORDER BY id ASC")
                self._write_to_excel(self.path_info, ["ID", "الاسم"], cur.fetchall(), "1A73E8")
                cur.execute(
                    "SELECT s.id, s.name, p.* FROM students s JOIN payments p ON s.id = p.sid ORDER BY s.id ASC")
                raw = cur.fetchall()
                clean_pays = [[r[0], r[1]] + list(r[3:]) for r in raw]
                self._write_to_excel(self.path_pay, ["ID", "الاسم"] + MONTHS, clean_pays, "34A853")
        except Exception as e:
            print(f"Error: {e}")

    def _write_to_excel(self, path, headers, data, color):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.append(headers)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF");
            cell.fill = fill;
            cell.alignment = Alignment(horizontal="center")
        for row in data: ws.append(row)
        wb.save(path)


class MainApp:
    def __init__(self, root, group):
        self.root = root
        self.data = DataEngine(group)
        self.root.title(f"نظام الإدارة الذكي - {group}")
        self.root.geometry("1000x850")
        self.root.configure(bg="#f4f7f6")
        self.create_widgets()
        self.refresh_table()

    def create_widgets(self):
        header = tk.Frame(self.root, bg="#2c3e50", pady=10)
        header.pack(fill="x")
        tk.Label(header, text=f"📂 {self.data.group}", font=("Arial", 14, "bold"), bg="#2c3e50", fg="white").pack(
            side="right", padx=20)
        tk.Button(header, text="📄 ملف الأسماء", bg="#3498db", fg="white",
                  command=lambda: os.startfile(self.data.path_info)).pack(side="left", padx=5)
        tk.Button(header, text="💰 سجل المدفوعات", bg="#27ae60", fg="white",
                  command=lambda: os.startfile(self.data.path_pay)).pack(side="left", padx=5)

        # إطار البحث الذكي (جديد)
        f_search = tk.LabelFrame(self.root, text=" 🔍 محرك البحث الذكي (بالاسم أو الـ ID) ", bg="white", padx=15,
                                 pady=10)
        f_search.pack(fill="x", padx=20, pady=10)

        tk.Label(f_search, text="ابحث هنا:", bg="white").pack(side="right")
        self.ent_search = tk.Entry(f_search, font=("Arial", 12), justify="right", width=30)
        self.ent_search.pack(side="right", padx=10)
        self.ent_search.bind("<KeyRelease>", self.smart_search)

        # شاشة عرض الحالة السريعة (تظهر عند البحث بالـ ID)
        self.lbl_info = tk.Label(f_search, text="ادخل ID لرؤية الحالة المالية فوراً", font=("Arial", 10, "bold"),
                                 fg="#7f8c8d", bg="#ecf0f1", width=50, pady=5)
        self.lbl_info.pack(side="left", padx=10)


        f_add = tk.LabelFrame(self.root, text=" إضافة طالب جديد ", bg="white", padx=15, pady=15)
        f_add.pack(fill="x", padx=20, pady=10)
        self.ent_name = tk.Entry(f_add, width=50, justify="right", font=("Arial", 12))
        self.ent_name.pack(side="right", padx=10)
        tk.Button(f_add, text="حفظ 💾", bg="#2ecc71", fg="white", font=("Arial", 10, "bold"),
                  command=self.save_student).pack(side="left")

        f_ops = tk.LabelFrame(self.root, text=" عمليات سداد وحذف ", bg="white", padx=15, pady=10)
        f_ops.pack(fill="x", padx=20, pady=10)

        tk.Label(f_ops, text="رقم الطالب:").pack(side="right")
        self.id_op = tk.Entry(f_ops, font=("Arial", 14, "bold"), width=8, justify="center")
        self.id_op.pack(side="right", padx=5)

        self.cb_month = ttk.Combobox(f_ops, values=MONTHS, state="readonly", width=12)
        self.cb_month.set(MONTHS[datetime.now().month - 1])
        self.cb_month.pack(side="right", padx=10)

        tk.Button(f_ops, text="سداد ✅", bg="#f1c40f", command=self.pay_id).pack(side="right", padx=5)
        tk.Button(f_ops, text="حذف 🗑️", bg="#e74c3c", fg="white", command=self.delete_student).pack(side="right",
                                                                                                    padx=5)

        # الجدول
        self.tree = ttk.Treeview(self.root, columns=("n", "i"), show="headings")
        self.tree.heading("n", text="اسم الطالب");
        self.tree.heading("i", text="ID")
        self.tree.column("i", width=100, anchor="center")
        self.tree.column("n", width=700, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=20, pady=10)

    def smart_search(self, event):
        val = self.ent_search.get().strip()
        month = self.cb_month.get()

        # إذا كان المدخل رقماً (بحث بالـ ID)
        if val.isdigit():
            with sqlite3.connect(self.data.db_path) as conn:
                cur = conn.cursor()
                cur.execute(
                    f"SELECT s.name, p.[{month}] FROM students s JOIN payments p ON s.id = p.sid WHERE s.id = ?",
                    (val,))
                res = cur.fetchone()
                if res:
                    status = "دفع ✅" if res[1] == '✅ مدفوع' else "لم يدفع ❌"
                    color = "#27ae60" if "✅" in status else "#e74c3c"
                    self.lbl_info.config(text=f"الطالب: {res[0]} | الحالة: {status}", fg="white", bg=color)
                else:
                    self.lbl_info.config(text="رقم غير موجود!", fg="white", bg="#95a5a6")
            self.refresh_table(query="", search_id=val)
        else:
            # بحث بالاسم
            self.lbl_info.config(text="ادخل ID لرؤية الحالة المالية فوراً", fg="#7f8c8d", bg="#ecf0f1")
            self.refresh_table(query=val)

    def save_student(self):
        n = self.ent_name.get().strip()
        if not n: return
        with sqlite3.connect(self.data.db_path) as conn:
            cur = conn.cursor()
            cur.execute("INSERT INTO students (name) VALUES (?)", (n,))
            sid = cur.lastrowid
            cur.execute("INSERT INTO payments (sid) VALUES (?)", (sid,))
        self.data.auto_save_all();
        self.refresh_table()
        self.ent_name.delete(0, 'end')

    def pay_id(self):
        sid, m = self.id_op.get().strip(), self.cb_month.get()
        if not sid: return
        with sqlite3.connect(self.data.db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT name FROM students WHERE id=?", (sid,))
            student = cur.fetchone()
            if student:
                conn.execute(f"UPDATE payments SET [{m}]='✅ مدفوع' WHERE sid=?", (sid,))
                self.data.auto_save_all();
                self.id_op.delete(0, 'end')
                messagebox.showinfo("سداد", f"تم تسجيل سداد {m} لـ {student[0]}")
                self.smart_search(None)  # تحديث لوحة الحالة فوراً
            else:
                messagebox.showerror("خطأ", "ID غير موجود")

    def delete_student(self):
        sid = self.id_op.get().strip()
        if sid and messagebox.askyesno("تأكيد", "حذف الطالب؟"):
            with sqlite3.connect(self.data.db_path) as conn:
                conn.execute("DELETE FROM students WHERE id=?", (sid,))
                conn.execute("DELETE FROM payments WHERE sid=?", (sid,))
            self.data.auto_save_all();
            self.refresh_table();
            self.id_op.delete(0, 'end')

    def refresh_table(self, query="", search_id=""):
        for i in self.tree.get_children(): self.tree.delete(i)
        with sqlite3.connect(self.data.db_path) as conn:
            if search_id:
                sql = "SELECT name, id FROM students WHERE id = ?"
                rows = conn.execute(sql, (search_id,))
            else:
                sql = "SELECT name, id FROM students WHERE name LIKE ? ORDER BY id ASC"
                rows = conn.execute(sql, (f"%{query}%",))
            for r in rows: self.tree.insert("", "end", values=r)


if __name__ == "__main__":
    def start():
        login = tk.Tk();
        login.title("نظام الإدارة");
        login.geometry("300x200")
        cb = ttk.Combobox(login, values=GROUPS, state="readonly");
        cb.set(GROUPS[-1]);
        cb.pack(pady=30)

        def go():
            g = cb.get();
            login.destroy();
            root = tk.Tk();
            MainApp(root, g);
            root.mainloop()

        tk.Button(login, text="دخول ✅", command=go, bg="#1a73e8", fg="white", width=15).pack()
        login.mainloop()


    start()